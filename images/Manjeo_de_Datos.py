from openpyxl import Workbook,load_workbook
from datetime import date,datetime
from views.Lista_de_productos import nombre_producto,Precio_producto,Stock_producto

hora_actual=datetime.now().strftime('%H:%M:%S')
fecha_actual=datetime.now().strftime("%d/%m/%Y")
wb=load_workbook("Base_de_Datos_Ventas.xlsx")
ventas=wb["Registro de ventas"]
productos=wb["Lista de Productos"]
encabezados=[cel.value for cel in productos[1]]
encabezados_ventas=[val.value for val in ventas[1]]
filas=[]
def poner_filas():
    wb= load_workbook("Base_de_Datos_Ventas.xlsx")
    productos= wb["Lista de Productos"]
    n_filas_productos = productos.max_row
    filas.clear()
    for i in range(2,n_filas_productos+1):
        fila=[celd.value for celd in productos[i]]
        filas.append(fila)  

def Obtener_num_venta():
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    n=f"{ventas.max_row:05d}"
    return n

def Obtener_num_producto():
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    productos=wb["Lista de Productos"]
    n2=f"{productos.max_row:05d}"
    return n2

def Registrar_venta(e):
    from views.Menu import tipo_pago
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    n=Obtener_num_venta()
    ventas.append([n,fecha_actual,hora_actual,mostrar_prod_sel(),tipo_pago.value,Cantidad_total(),Total()])
    wb.save("Base_de_Datos_Ventas.xlsx")

def Registrar_producto(e):
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    productos=wb["Lista de Productos"]
    n2=f"{productos.max_row:05d}"
    productos.append([n2,nombre_producto.value,Stock_producto.value,Precio_producto.value])  
    nombre_producto.value=""
    Precio_producto.value=""
    Stock_producto.value=""

    wb.save("Base_de_Datos_Ventas.xlsx")

def poner_filas_producto():
    import flet as ft
    listas=[]
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    productos=wb["Lista de Productos"]
    for i in range(2,productos.max_row+1):
        lista=[ft.Text(productos[f"B{i}"].value),ft.Text(productos[f"D{i}"].value),ft.TextField(width=60),ft.Checkbox()]
        listas.append(lista)
    
    return listas

def poner_filas_vent():
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    filas_ventas=[]
    for i in range(2,ventas.max_row+1):
        filas=[vent.value for vent in ventas[i]]
        filas_ventas.append(filas)
    
    return filas_ventas

        
def Total():
    from views.Menu import cantidades,precio_venta
    suma=0
    a=0
    for i in cantidades:
        
        multiplicacion=float(cantidades[a])*float(precio_venta[a])
        suma+=multiplicacion
        a+=1
    
    return suma

def mostrar_prod_sel():
    from views.Menu import nom_prod_selec, cantidades
    produc=[]
    for i in range(len(cantidades)):
        produc.append(f"{nom_prod_selec[i]}({cantidades[i]})")
    prod_con_cant=", ".join(produc)
    return prod_con_cant

def Cantidad_total():
    from views.Menu import cantidades
    cantidad_total=0
    for i in cantidades:
        
        cantidad_total+=int(i)

    return cantidad_total    

def actualizar_tabla_ventas(mes,año):
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    filas_de_x_mes=[]
    if año==None:
        año=str(datetime.now().year)
    
    if mes==None:
        mes="Todos"
    

    for i in range(2,ventas.max_row+1):
        
        if str(ventas[i][1].value)[3:5]==mes and str(ventas[i][1].value)[6:10]==año:
            filas=[celd.value for celd in ventas[i]]
            filas_de_x_mes.append(filas)
        elif "Todos"==mes and str(ventas[i][1].value)[6:10]==año:
            filas=[celd.value for celd in ventas[i]]
            filas_de_x_mes.append(filas)
        
        

    return filas_de_x_mes
meses=["0","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]
def Generar_reporte(mes_inicio,mes_fin,año_inicio,año_fin):
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    
    mes_inicio_l=meses[int(mes_inicio)]
    mes_finl=meses[int(mes_fin)]
    periodo=f"Periodo:Desde {mes_inicio_l} {año_inicio} hasta {mes_finl} {año_fin} "
    datos_calculados = sacar_datos(mes_inicio, mes_fin, año_inicio, año_fin)
    Cant_ventas=f"Cantidad de ventas: {datos_calculados[0]}"
    Cant_prod_vend=f"Cantidad de productos vendidos: {datos_calculados[2]}"
    Total_vendid=f"Total S/ {datos_calculados[1]}"
    Informacion=[periodo,Cant_ventas,Cant_prod_vend,Total_vendid]
    return Informacion

def sacar_datos(mes_inicio,mes_fin,año_inicio,año_fin):
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    cant_ventas=0
    Total_vendid=0
    Cant_prod_vend=0
    meses_validos_inicio=[f"{a:02d}" for a in range(int(mes_inicio),13)]
    meses_validos_fin=[f"{a:02d}" for a in range(1,int(mes_fin)+1)]
    meses_validos_igual=[f"{a:02d}" for a in range(int(mes_inicio),int(mes_fin)+1)]
    años_validos=[str(b) for b in range(int(año_inicio),int(año_fin)+1)]
    print(meses_validos_inicio)
    print(meses_validos_fin)
    print(años_validos)

    for i in range(2,ventas.max_row+1):
        if int(año_fin)==int(año_inicio):
            if str(ventas[i][1].value)[6:10] in años_validos and str(ventas[i][1].value)[3:5] in meses_validos_igual:
                fila=[a.value for a in ventas[i]]
                Cant_prod_vend+=int(fila[5])
                Total_vendid+=int(fila[6])
                cant_ventas+=1
        else:
            if str(ventas[i][1].value)[6:10] in años_validos:
                if str(ventas[i][1].value)[3:5] in meses_validos_inicio and str(ventas[i][1].value)[6:10]==años_validos[0]:
                    fila=[a.value for a in ventas[i]]
                    Cant_prod_vend+=int(fila[5])
                    Total_vendid+=int(fila[6])
                    cant_ventas+=1
                elif str(ventas[i][1].value)[6:10]!=años_validos[0] and str(ventas[i][1].value)[6:10]!=años_validos[-1]:
                    fila=[a.value for a in ventas[i]]
                    Cant_prod_vend+=int(fila[5])
                    Total_vendid+=int(fila[6])
                    cant_ventas+=1
                elif str(ventas[i][1].value)[3:5] in meses_validos_fin and str(ventas[i][1].value)[6:10]==años_validos[-1]:
                    fila=[a.value for a in ventas[i]]
                    Cant_prod_vend+=int(fila[5])
                    Total_vendid+=int(fila[6])
                    cant_ventas+=1
            

    return [cant_ventas, Total_vendid,Cant_prod_vend]


def detalle_por_año(año):
    from views.Reporte import desplegable_año,desplegable_año2,desplegable_mes,desplegable_mes2
    wb=load_workbook("Base_de_Datos_Ventas.xlsx")
    ventas=wb["Registro de ventas"]
    cant_ventas=[]
    Total_vendid=[]
    Cant_prod_vend=[]
    cv_meses=[0,0,0,0,0,0,0,0,0,0,0,0,0]
    cpv_meses=[0,0,0,0,0,0,0,0,0,0,0,0,0]
    Ctv_meses=[0,0,0,0,0,0,0,0,0,0,0,0,0]
    for i in range(2,ventas.max_row+1):
        if str(ventas[i][1].value)[6:10]==str(año):
            for j in range(1,13):
                if f"{j:02d}"==str(ventas[i][1].value)[3:5]:
                    cv_meses[j]+=1
                    cpv_meses[j]+=int(ventas[i][5].value)
                    Ctv_meses[j]+=int(ventas[i][6].value)
    
    for h in range(1,13):
                        cant_ventas.append(cv_meses[int(h)])
                        Cant_prod_vend.append(cpv_meses[int(h)])
                        Total_vendid.append(Ctv_meses[int(h)])
    return [cant_ventas, Total_vendid,Cant_prod_vend]